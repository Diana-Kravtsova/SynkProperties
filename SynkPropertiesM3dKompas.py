"""
Синхронизатор свойств КОМПАС-3D
================================
Сканирует сборочный файл (.a3d), сравнивает свойства компонентов
(обозначение, наименование, примечание) между файлом детали и записью
в сборочнике/спецификации. Для штоков и гильз дополнительно проверяет
соответствие числа длины в обозначении и в поле примечания.
Позволяет применить исправления прямо из GUI и выгрузить отчёт в Excel.
"""
import os
import re
import queue
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk

import pythoncom
import win32com.client

LIBRARY_PATHS = [
    r"C:\Program Files\ASCON",
    r"C:\ProgramData\ASCON",
]

# ID стандартных свойств КОМПАС
PROP_ID_MARKING = 4.0
PROP_ID_NAME = 5.0
PROP_ID_COMMENT_FALLBACK = 7.0

# === КОНФИГУРАЦИЯ ТИПОВ ДЕТАЛЕЙ ===
# Ключ = префикс обозначения (регистронезависимо)
# value = позиция числа длины в обозначении (0-based), описание
PART_TYPE_CONFIG = {
    "RD": {"length_position": 2, "description": "Шток"},   
    "TB": {"length_position": 3, "description": "Гильза"},    
}


def _find_property_by_name(property_mng, doc, search_names):
    """Ищет свойство документа по одному из допустимых имён (регистронезависимо).
    Возвращает первое совпадение или None.
    """
    # Строим множество один раз, а не на каждой итерации цикла
    names_lower = {n.lower() for n in search_names}
    try:
        props = property_mng.GetProperties(doc, 0)
        for i in range(props.Count):
            prop = props.Item(i)
            if (prop.Name or "").strip().lower() in names_lower:
                return prop
    except Exception:
        pass
    return None


def _detect_part_type(marking):
    """Определяет тип детали по префиксу обозначения"""
    if not marking:
        return None
    prefix = marking.split('.')[0].upper() if '.' in marking else marking[:2].upper()
    return PART_TYPE_CONFIG.get(prefix)


def _extract_length_from_marking(marking, part_type=None):
    """Извлекает длину из обозначения с учётом типа детали"""
    if not marking:
        return None

    parts = marking.split('.')

    # Если тип определён — берём позицию из конфига
    if part_type:
        pos = part_type.get("length_position", 2)
        if pos < len(parts):
            clean = re.sub(r'[()A-Za-z]', '', parts[pos])
            # Убираем ведущие нули
            clean = clean.lstrip('0')
            if clean.isdigit() and len(clean) >= 1:
                return int(clean)
        return None

    # Fallback: ищем число >= 2 цифр в любом месте
    for part in parts:
        clean = re.sub(r'[()A-Za-z]', '', part)
        clean = clean.lstrip('0')
        if clean.isdigit() and len(clean) >= 2:
            return int(clean)
    return None


def _extract_length_from_comment(comment):
    """Извлекает числовое значение длины из строки примечания.
    Приоритет: явный формат 'L = 312', иначе — первое двузначное+ число.
    """
    if not comment:
        return None
    match = re.search(r'L\s*=\s*(\d+)', comment, re.IGNORECASE)
    if match:
        return int(match.group(1))
    match = re.search(r'\b(\d{2,})\b', comment)
    if match:
        return int(match.group(1))
    return None


class KompasSyncApp(ctk.CTk):

    def __init__(self):
        super().__init__()
        self.title("Синхронизатор свойств КОМПАС-3D v24")
        self.geometry("1700x950")
        self.asm_path = ""
        self.conflicts = []
        self.all_parts = []
        self.queue = queue.Queue()
        self.stop_requested = False
        self.init_ui()
        self.check_queue()

    def init_ui(self):
        self.file_frame = ctk.CTkFrame(self)
        self.file_frame.pack(padx=20, pady=15, fill="x")

        self.btn_browse = ctk.CTkButton(
            self.file_frame, text="Выбрать сборку (.a3d)", command=self.browse_file
        )
        self.btn_browse.pack(side="left", padx=10, pady=10)

        self.lbl_file = ctk.CTkLabel(
            self.file_frame, text="Файл не выбран", text_color="gray"
        )
        self.lbl_file.pack(side="left", padx=10, pady=10)

        self.btn_scan = ctk.CTkButton(
            self.file_frame, text="Сканировать", state="disabled", command=self.start_scan
        )
        self.btn_scan.pack(side="right", padx=10, pady=10)

        self.status_frame = ctk.CTkFrame(self)
        self.status_frame.pack(padx=20, pady=5, fill="x")

        self.lbl_status = ctk.CTkLabel(
            self.status_frame, text="Выберите сборку для начала работы", font=("Arial", 14)
        )
        self.lbl_status.pack(side="left", padx=10, pady=5)

        self.btn_stop = ctk.CTkButton(
            self.status_frame, text="⏹ Остановить", state="disabled",
            command=self.stop_scan, width=120, fg_color="red", hover_color="darkred"
        )
        self.btn_stop.pack(side="right", padx=5, pady=5)

        self.btn_export = ctk.CTkButton(
            self.status_frame, text="📊 Экспорт в Excel", state="disabled",
            command=self.export_to_excel, width=140
        )
        self.btn_export.pack(side="right", padx=5, pady=5)

        self.btn_select_all_src = ctk.CTkButton(
            self.status_frame, text="☑ Всё из детали", state="disabled",
            command=lambda: self.mass_select("pull_from_src"), width=140
        )
        self.btn_select_all_src.pack(side="right", padx=5, pady=5)

        self.btn_select_all_leave = ctk.CTkButton(
            self.status_frame, text="☐ Снять все", state="disabled",
            command=lambda: self.mass_select("leave"), width=120
        )
        self.btn_select_all_leave.pack(side="right", padx=5, pady=5)

        self.scroll_frame = ctk.CTkScrollableFrame(self, label_text="Найденные конфликты свойств")
        self.scroll_frame.pack(padx=20, pady=10, fill="both", expand=True)

        self.bottom_frame = ctk.CTkFrame(self)
        self.bottom_frame.pack(padx=20, pady=15, fill="x")

        self.progress = ctk.CTkProgressBar(self.bottom_frame, width=300)
        self.progress.pack(side="left", padx=10, pady=10)
        self.progress.set(0)

        self.btn_apply = ctk.CTkButton(
            self.bottom_frame, text="Применить выбранные изменения",
            state="disabled", command=self.apply_changes
        )
        self.btn_apply.pack(side="right", padx=10, pady=10)

    def is_library_part(self, filepath):
        filepath_lower = filepath.lower()
        for lib_path in LIBRARY_PATHS:
            if lib_path.lower() in filepath_lower:
                return True
        return False

    def browse_file(self):
        file = filedialog.askopenfilename(filetypes=[("Сборки КОМПАС-3D", "*.a3d")])
        if file:
            self.asm_path = os.path.normpath(file)
            self.lbl_file.configure(text=os.path.basename(self.asm_path), text_color="white")
            self.btn_scan.configure(state="normal")

    def stop_scan(self):
        self.stop_requested = True
        self.lbl_status.configure(text="Остановка...")
        self.btn_stop.configure(state="disabled")

    def start_scan(self):
        self.stop_requested = False
        self.btn_scan.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.lbl_status.configure(text="Сканирование сборки...")
        self.progress.set(0)

        for widget in self.scroll_frame.winfo_children():
            widget.destroy()

        self.btn_select_all_src.configure(state="disabled")
        self.btn_select_all_leave.configure(state="disabled")
        self.btn_apply.configure(state="disabled")
        self.btn_export.configure(state="disabled")

        threading.Thread(
            target=self.scan_kompas_file_thread, args=(self.asm_path,), daemon=True
        ).start()

    def _get_property_value(self, property_keeper, property_mng, doc, prop_id, prop_names=None):
        try:
            if prop_id:
                prop = property_mng.GetProperty(doc, prop_id)
                if prop:
                    result = property_keeper.GetPropertyValue(prop, '', True, True)
                    if result and len(result) > 1 and result[1]:
                        return (result[1] or "").strip(), prop_id

            if prop_names:
                prop = _find_property_by_name(property_mng, doc, prop_names)
                if prop:
                    result = property_keeper.GetPropertyValue(prop, '', True, True)
                    if result and len(result) > 1 and result[1]:
                        return (result[1] or "").strip(), prop.Id

            return "", prop_id
        except Exception as e:
            print(f"  Ошибка получения свойства ID={prop_id}, names={prop_names}: {e}")
            return "", prop_id

    def _set_property_value(self, property_keeper, property_mng, doc, prop_id, value, prop_names=None):
        try:
            prop = None
            if prop_id:
                prop = property_mng.GetProperty(doc, prop_id)
            if not prop and prop_names:
                prop = _find_property_by_name(property_mng, doc, prop_names)
            if prop:
                property_keeper.SetPropertyValue(prop, value, True)
                return True
            return False
        except Exception as e:
            print(f"  Ошибка установки свойства ID={prop_id}: {e}")
            return False

    def _get_properties_from_part(self, part, property_mng, doc, comment_id=None):
        """Читает наименование, обозначение и примечание из IPropertyKeeper компонента.
        comment_id кешируется между вызовами — при первом успешном чтении
        ID свойства запоминается и используется напрямую в следующих итерациях.
        """
        try:
            keeper = win32com.client.CastTo(part, "IPropertyKeeper")
            if not keeper:
                return {'name': '', 'marking': '', 'comment': ''}, comment_id

            name_val, _ = self._get_property_value(keeper, property_mng, doc, PROP_ID_NAME)
            marking_val, _ = self._get_property_value(keeper, property_mng, doc, PROP_ID_MARKING)
            comment_val, found_id = self._get_property_value(
                keeper, property_mng, doc,
                comment_id or PROP_ID_COMMENT_FALLBACK,
                prop_names=["Примечание", "Comment", "Комментарий"]
            )
            # Запоминаем ID свойства-примечания после первого успешного определения
            if found_id and not comment_id:
                comment_id = found_id

            return {
                'name': name_val,
                'marking': marking_val,
                'comment': comment_val,
            }, comment_id
        except Exception as e:
            print(f"  Ошибка получения свойств part: {e}")
            return {'name': '', 'marking': '', 'comment': ''}, comment_id

    def _get_properties_from_spec(self, obj):
        """Читает свойства компонента из строки спецификации.
        ColumnType: 4 = обозначение, 5 = наименование, 6 = примечание.
        """
        try:
            cols = obj.Columns
            result = {'marking': '', 'name': '', 'comment': ''}
            for j in range(cols.Count):
                col = cols.Item(j)
                col_type = str(col.ColumnType)
                items = col.ColumnItems
                if items and items.Count > 0:
                    val = (items.Item(0).Value or "").strip()
                    if col_type == '4':
                        result['marking'] = val
                    elif col_type == '5':
                        result['name'] = val
                    elif col_type == '6':
                        result['comment'] = val
            return result
        except Exception:
            return {'marking': '', 'name': '', 'comment': ''}

    def _get_expected_filename(self, marking, name):
        marking = marking.strip()
        name = name.strip()
        if marking and name:
            return f"{marking} {name}"
        elif marking:
            return marking
        elif name:
            return name
        return ""

    def scan_kompas_file_thread(self, filepath):
        pythoncom.CoInitialize()
        try:
            kompas_api = win32com.client.Dispatch("KOMPAS.Application.7")
            app = kompas_api.Application

            try:
                property_mng = win32com.client.CastTo(app, "IPropertyMng")
            except Exception:
                property_mng = None
                print("WARNING: Не удалось получить IPropertyMng")

            doc = app.Documents.Open(filepath, True, False)
            if not doc:
                self.queue.put(("error", "Не удалось открыть файл сборки."))
                return

            doc3d = win32com.client.CastTo(doc, "IKompasDocument3D")
            top_part = doc3d.TopPart
            parts = top_part.Parts

            spec = doc3d.SpecificationDescriptions
            has_spec = spec.Count > 0
            spec_desc = spec.Item(0) if has_spec else None
            base = spec_desc.BaseObjects if has_spec else None

            found_conflicts = []
            self.all_parts = []
            total = parts.Count
            comment_id = None

            for i in range(total):
                if self.stop_requested:
                    break

                self.queue.put(("progress", (i + 1) / total * 0.9))

                try:
                    comp = parts.Part(i)
                except Exception:
                    continue

                source_path = getattr(comp, 'FileName', '') or ''
                if not source_path:
                    continue

                if self.is_library_part(source_path):
                    continue

                if not os.path.isabs(source_path):
                    source_path = os.path.normpath(os.path.join(os.path.dirname(filepath), source_path))

                if not os.path.exists(source_path):
                    continue

                comp_ref = getattr(comp, 'Reference', None)

                if property_mng:
                    comp_props, comment_id = self._get_properties_from_part(comp, property_mng, doc, comment_id)
                else:
                    comp_props = {'name': (getattr(comp, 'Name', '') or '').strip(),
                                  'marking': (getattr(comp, 'Marking', '') or '').strip(),
                                  'comment': ''}

                spec_props = None
                if has_spec and base:
                    try:
                        for j in range(base.Count):
                            obj = base.Item(j)
                            geom = obj.Geometry[0]
                            if geom.Reference == comp_ref:
                                spec_props = self._get_properties_from_spec(obj)
                                break
                    except Exception:
                        pass

                if spec_props and (spec_props['name'] or spec_props['marking'] or spec_props['comment']):
                    asm_props = spec_props
                else:
                    asm_props = comp_props

                src_doc = None
                src_props = {'name': '', 'marking': '', 'comment': ''}
                try:
                    src_doc = app.Documents.Open(source_path, False, True)
                    if src_doc:
                        src_doc3d = win32com.client.CastTo(src_doc, "IKompasDocument3D")
                        src_part = src_doc3d.TopPart
                        if src_part and property_mng:
                            src_props, comment_id = self._get_properties_from_part(src_part, property_mng, src_doc, comment_id)
                        elif src_part:
                            src_props = {
                                'name': (getattr(src_part, 'Name', '') or '').strip(),
                                'marking': (getattr(src_part, 'Marking', '') or '').strip(),
                                'comment': ''
                            }
                        src_doc.Close(2)
                except Exception as e:
                    print(f"Ошибка открытия детали {source_path}: {e}")
                    if src_doc:
                        try:
                            src_doc.Close(2)
                        except Exception:
                            pass

                # === ОПРЕДЕЛЯЕМ ТИП ДЕТАЛИ ===
                part_type = _detect_part_type(src_props['marking'])
                part_type_name = part_type["description"] if part_type else ""

                # === ПРОВЕРКА ДЛИНЫ (только для штока и гильзы) ===
                marking_len = None
                comment_len = None
                length_mismatch = False

                if part_type:  
                    marking_len = _extract_length_from_marking(src_props['marking'], part_type)
                    comment_len = _extract_length_from_comment(src_props['comment'])
                    if marking_len and comment_len and marking_len != comment_len:
                        length_mismatch = True

                # Проверка имени файла
                actual_name = os.path.splitext(os.path.basename(source_path))[0]
                expected_name = self._get_expected_filename(src_props['marking'], src_props['name'])
                filename_mismatch = (actual_name != expected_name) and expected_name

                self.all_parts.append({
                    'file_path': source_path,
                    'display_name': os.path.basename(source_path),
                    'part_type': part_type_name,
                    'src_marking': src_props['marking'],
                    'src_name': src_props['name'],
                    'src_comment': src_props['comment'],
                    'asm_marking': asm_props['marking'],
                    'asm_name': asm_props['name'],
                    'asm_comment': asm_props['comment'],
                    'expected_filename': expected_name,
                    'actual_filename': actual_name,
                    'marking_length': marking_len,
                    'comment_length': comment_len,
                    'length_mismatch': length_mismatch,
                })

                diffs = []
                if asm_props['name'] != src_props['name']:
                    diffs.append("Наименование")
                if asm_props['marking'] != src_props['marking']:
                    diffs.append("Обозначение")
                if asm_props['comment'] != src_props['comment']:
                    diffs.append("Примечание")
                if filename_mismatch:
                    diffs.append("Имя файла")
                if length_mismatch:
                    diffs.append("Длина")

                if diffs:
                    found_conflicts.append({
                        "comp_index": i,
                        "comp_ref": comp_ref,
                        "file_path": source_path,
                        "display_name": os.path.basename(source_path),
                        "part_type": part_type_name,
                        "asm_props": asm_props,
                        "src_props": src_props,
                        "diffs": diffs,
                        "marking_len": marking_len,
                        "comment_len": comment_len,
                        "length_mismatch": length_mismatch,
                        "decision": tk.StringVar(value="leave"),
                    })

            doc.Close(2)

            if self.stop_requested:
                self.queue.put(("scan_done", []))
            else:
                self.queue.put(("scan_done", found_conflicts))

        except Exception as e:
            import traceback
            self.queue.put(("error", f"Ошибка API: {str(e)}\n\n{traceback.format_exc()}"))
        finally:
            pythoncom.CoUninitialize()

    def check_queue(self):
        try:
            while True:
                msg_type, data = self.queue.get_nowait()
                if msg_type == "scan_done":
                    self.conflicts = data
                    self.lbl_status.configure(text=f"Сканирование завершено. Конфликтов: {len(data)} из {len(self.all_parts)}")
                    self.build_conflict_ui()
                    self.btn_scan.configure(state="normal")
                    self.btn_stop.configure(state="disabled")
                    self.progress.set(1)
                    if data:
                        self.btn_apply.configure(state="normal")
                        self.btn_select_all_src.configure(state="normal")
                        self.btn_select_all_leave.configure(state="normal")
                    if self.all_parts:
                        self.btn_export.configure(state="normal")
                elif msg_type == "apply_done":
                    messagebox.showinfo("Готово", f"Изменения применены: {data} объектов.")
                    self.lbl_status.configure(text=f"Изменения применены: {data} объектов.")
                    self.btn_apply.configure(state="normal")
                    self.progress.set(1)
                    self.start_scan()
                elif msg_type == "error":
                    messagebox.showerror("Ошибка", data)
                    self.lbl_status.configure(text="Произошла ошибка.")
                    self.btn_scan.configure(state="normal")
                    self.btn_stop.configure(state="disabled")
                    self.btn_apply.configure(state="normal")
                    self.progress.set(0)
                elif msg_type == "progress":
                    self.progress.set(data)
                elif msg_type == "apply_progress":
                    self.progress.set(data)
        except queue.Empty:
            pass
        finally:
            self.after(100, self.check_queue)

    def mass_select(self, value):
        for conflict in self.conflicts:
            conflict["decision"].set(value)

    def _create_table_row(self, parent, row, col_data, is_header=False):
        bg_color = "#2B2B2B" if not is_header else "#366092"
        font = ("Consolas", 11, "bold") if is_header else ("Consolas", 10)
        for col_idx, (text, width, fg_color) in enumerate(col_data):
            lbl = ctk.CTkLabel(
                parent, text=text, width=width, anchor="w",
                font=font, text_color=fg_color,
                fg_color=bg_color, corner_radius=0
            )
            lbl.grid(row=row, column=col_idx, padx=1, pady=1, sticky="ew")

    def build_conflict_ui(self):
        if not self.conflicts:
            ctk.CTkLabel(self.scroll_frame, text="Конфликтов не обнаружено. Всё синхронно!").pack(pady=20)
            return

        for idx, conflict in enumerate(self.conflicts):
            row_frame = ctk.CTkFrame(self.scroll_frame)
            row_frame.pack(fill="x", padx=5, pady=8)

            # Заголовок с типом детали
            type_label = f" [{conflict['part_type']}]" if conflict['part_type'] else ""
            header = ctk.CTkLabel(
                row_frame, text=f"📁 {conflict['display_name']}{type_label}",
                font=("Arial", 13, "bold"), anchor="w"
            )
            header.pack(fill="x", padx=10, pady=(5, 0))

            table_frame = ctk.CTkFrame(row_frame, fg_color="transparent")
            table_frame.pack(fill="x", padx=10, pady=5)

            headers = [
                ("Свойство", 120, "white"),
                ("Сборочник", 300, "white"),
                ("Деталь", 300, "white"),
                ("Статус", 80, "white"),
            ]
            self._create_table_row(table_frame, 0, headers, is_header=True)

            src = conflict['src_props']
            asm = conflict['asm_props']
            actual_name = os.path.splitext(os.path.basename(conflict['file_path']))[0]
            expected_name = self._get_expected_filename(src['marking'], src['name'])

            # Обозначение
            marking_match = asm['marking'] == src['marking']
            marking_data = [
                ("Обозначение", 120, "#AAAAAA"),
                (asm['marking'] or "—", 300, "#4CAF50" if marking_match else "#FF6B6B"),
                (src['marking'] or "—", 300, "#4CAF50" if marking_match else "#FFB347"),
                ("✓" if marking_match else "✗", 80, "#4CAF50" if marking_match else "#FF4444"),
            ]
            self._create_table_row(table_frame, 1, marking_data)

            # Наименование
            name_match = asm['name'] == src['name']
            name_data = [
                ("Наименование", 120, "#AAAAAA"),
                (asm['name'] or "—", 300, "#4CAF50" if name_match else "#FF6B6B"),
                (src['name'] or "—", 300, "#4CAF50" if name_match else "#FFB347"),
                ("✓" if name_match else "✗", 80, "#4CAF50" if name_match else "#FF4444"),
            ]
            self._create_table_row(table_frame, 2, name_data)

            # Примечание
            comment_match = asm['comment'] == src['comment']
            comment_data = [
                ("Примечание", 120, "#AAAAAA"),
                (asm['comment'] or "—", 300, "#4CAF50" if comment_match else "#FF6B6B"),
                (src['comment'] or "—", 300, "#4CAF50" if comment_match else "#FFB347"),
                ("✓" if comment_match else "✗", 80, "#4CAF50" if comment_match else "#FF4444"),
            ]
            self._create_table_row(table_frame, 3, comment_data)

            # Имя файла
            filename_match = actual_name == expected_name
            filename_data = [
                ("Имя файла", 120, "#AAAAAA"),
                (actual_name, 300, "#4CAF50" if filename_match else "#FFD700"),
                (expected_name or "—", 300, "#AAAAAA"),
                ("✓" if filename_match else "⚠", 80, "#4CAF50" if filename_match else "#FFD700"),
            ]
            self._create_table_row(table_frame, 4, filename_data)

            # Длина (только для штока/гильзы)
            if conflict.get('length_mismatch'):
                len_data = [
                    ("Длина", 120, "#FF6B6B"),
                    (f"Из обозн: {conflict['marking_len']}", 300, "#FF6B6B"),
                    (f"Из прим: {conflict['comment_len']}", 300, "#FF6B6B"),
                    ("✗", 80, "#FF4444"),
                ]
                self._create_table_row(table_frame, 5, len_data)
            elif conflict.get('marking_len') and conflict.get('comment_len'):
                len_data = [
                    ("Длина", 120, "#AAAAAA"),
                    (f"Из обозн: {conflict['marking_len']}", 300, "#4CAF50"),
                    (f"Из прим: {conflict['comment_len']}", 300, "#4CAF50"),
                    ("✓", 80, "#4CAF50"),
                ]
                self._create_table_row(table_frame, 5, len_data)

            radio_frame = ctk.CTkFrame(row_frame, fg_color="transparent")
            radio_frame.pack(fill="x", padx=10, pady=(0, 5))

            ctk.CTkRadioButton(
                radio_frame, text="Обновить из детали",
                variable=conflict["decision"], value="pull_from_src"
            ).pack(side="left", padx=10)

            ctk.CTkRadioButton(
                radio_frame, text="Не менять",
                variable=conflict["decision"], value="leave"
            ).pack(side="left", padx=10)

    def apply_changes(self):
        to_apply = [c for c in self.conflicts if c["decision"].get() == "pull_from_src"]
        if not to_apply:
            messagebox.showinfo("Информация", "Ничего не выбрано для изменения.")
            return

        self.btn_apply.configure(state="disabled")
        self.lbl_status.configure(text="Применение изменений...")
        self.progress.set(0.1)

        threading.Thread(
            target=self.apply_changes_thread, args=(to_apply, self.asm_path), daemon=True
        ).start()

    def apply_changes_thread(self, conflicts_to_apply, asm_path):
        pythoncom.CoInitialize()
        applied_count = 0

        try:
            kompas_api = win32com.client.Dispatch("KOMPAS.Application.7")
            app = kompas_api.Application

            try:
                property_mng = win32com.client.CastTo(app, "IPropertyMng")
            except Exception:
                property_mng = None

            doc = None
            norm_asm = os.path.normpath(asm_path)
            for d in app.Documents:
                try:
                    if os.path.normpath(d.PathName) == norm_asm:
                        doc = d
                        break
                except Exception:
                    continue

            if not doc:
                doc = app.Documents.Open(asm_path, True, False)

            if not doc:
                self.queue.put(("error", "Не удалось открыть сборку."))
                return

            doc3d = win32com.client.CastTo(doc, "IKompasDocument3D")
            top_part = doc3d.TopPart
            parts = top_part.Parts

            comp_map = {}
            for i in range(parts.Count):
                try:
                    c = parts.Part(i)
                    ref = getattr(c, 'Reference', None)
                    if ref:
                        comp_map[ref] = c
                except Exception:
                    continue

            spec = doc3d.SpecificationDescriptions
            has_spec = spec.Count > 0
            spec_desc = spec.Item(0) if has_spec else None
            base = spec_desc.BaseObjects if has_spec else None

            total = len(conflicts_to_apply)

            for idx, conflict in enumerate(conflicts_to_apply):
                self.queue.put(("apply_progress", 0.1 + (idx / total) * 0.8))
                print(f"\n--- Обработка: {conflict['display_name']} ---")

                comp = comp_map.get(conflict.get("comp_ref"))
                if not comp:
                    try:
                        comp = parts.Part(conflict["comp_index"])
                    except Exception:
                        continue

                if property_mng:
                    try:
                        keeper = win32com.client.CastTo(comp, "IPropertyKeeper")
                        if keeper:
                            self._set_property_value(keeper, property_mng, doc, PROP_ID_NAME,
                                conflict["src_props"]["name"])
                            self._set_property_value(keeper, property_mng, doc, PROP_ID_MARKING,
                                conflict["src_props"]["marking"])
                            self._set_property_value(keeper, property_mng, doc, PROP_ID_COMMENT_FALLBACK,
                                conflict["src_props"]["comment"],
                                prop_names=["Примечание", "Comment", "Комментарий"])
                            applied_count += 1
                    except Exception as e:
                        print(f"  ОШИБКА обновления свойств: {e}")

                if has_spec and base and conflict.get("comp_ref"):
                    try:
                        for j in range(base.Count):
                            obj = base.Item(j)
                            geom = obj.Geometry[0]
                            if geom.Reference == conflict["comp_ref"]:
                                cols = obj.Columns
                                for k in range(cols.Count):
                                    col = cols.Item(k)
                                    col_type = col.ColumnType
                                    items = col.ColumnItems
                                    if items and items.Count > 0:
                                        if col_type == '4':
                                            items.Item(0).Value = conflict["src_props"]["marking"]
                                        elif col_type == '5':
                                            items.Item(0).Value = conflict["src_props"]["name"]
                                        elif col_type == '6':
                                            items.Item(0).Value = conflict["src_props"]["comment"]
                                obj.SyncronizeWithProperties = True
                                obj.Update()
                                break
                    except Exception as e:
                        print(f"  ОШИБКА обновления спецификации: {e}")

            doc.Save()
            try:
                doc.Close(0)
            except Exception:
                pass

            self.queue.put(("apply_done", applied_count))

        except Exception as e:
            import traceback
            self.queue.put(("error", f"Ошибка применения: {str(e)}\n\n{traceback.format_exc()}"))
        finally:
            pythoncom.CoUninitialize()

    def export_to_excel(self):
        if not self.all_parts:
            messagebox.showinfo("Информация", "Нет данных для экспорта.")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Входящие детали"

            # Стили создаём один раз — переиспользуем для всех строк
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            center_align = Alignment(vertical="center")

            headers = [
                "№", "Тип", "Имя файла",
                "Обозначение (деталь)", "Наименование (деталь)", "Примечание (деталь)",
                "Обозначение (сборка)", "Наименование (сборка)", "Примечание (сборка)",
                "Длина (обозн)", "Длина (прим)", "Совпадение длины",
                "Ожидаемое имя файла", "Фактическое имя файла", "Совпадение имени"
            ]

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            for row_idx, part in enumerate(self.all_parts, 2):
                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                ws.cell(row=row_idx, column=2, value=part['part_type'])
                ws.cell(row=row_idx, column=3, value=part['display_name'])
                ws.cell(row=row_idx, column=4, value=part['src_marking'])
                ws.cell(row=row_idx, column=5, value=part['src_name'])
                ws.cell(row=row_idx, column=6, value=part['src_comment'])
                ws.cell(row=row_idx, column=7, value=part['asm_marking'])
                ws.cell(row=row_idx, column=8, value=part['asm_name'])
                ws.cell(row=row_idx, column=9, value=part['asm_comment'])
                ws.cell(row=row_idx, column=10, value=part['marking_length'] if part['marking_length'] else "")
                ws.cell(row=row_idx, column=11, value=part['comment_length'] if part['comment_length'] else "")

                if part['marking_length'] and part['comment_length']:
                    len_match = "Да" if not part['length_mismatch'] else "Нет"
                else:
                    len_match = "Н/Д"
                ws.cell(row=row_idx, column=12, value=len_match)

                ws.cell(row=row_idx, column=13, value=part['expected_filename'])
                ws.cell(row=row_idx, column=14, value=part['actual_filename'])

                name_match = "Да" if part['actual_filename'] == part['expected_filename'] else "Нет"
                ws.cell(row=row_idx, column=15, value=name_match)

                for col in range(1, 16):
                    ws.cell(row=row_idx, column=col).border = thin_border
                    ws.cell(row=row_idx, column=col).alignment = center_align

                if part['src_marking'] != part['asm_marking']:
                    ws.cell(row=row_idx, column=4).fill = red_fill
                    ws.cell(row=row_idx, column=7).fill = red_fill
                if part['src_name'] != part['asm_name']:
                    ws.cell(row=row_idx, column=5).fill = red_fill
                    ws.cell(row=row_idx, column=8).fill = red_fill
                if part['src_comment'] != part['asm_comment']:
                    ws.cell(row=row_idx, column=6).fill = red_fill
                    ws.cell(row=row_idx, column=9).fill = red_fill

                if part['length_mismatch']:
                    ws.cell(row=row_idx, column=10).fill = red_fill
                    ws.cell(row=row_idx, column=11).fill = red_fill
                    ws.cell(row=row_idx, column=12).fill = red_fill
                elif part['marking_length'] and part['comment_length']:
                    ws.cell(row=row_idx, column=12).fill = green_fill

                if name_match == "Нет":
                    ws.cell(row=row_idx, column=14).fill = yellow_fill
                    ws.cell(row=row_idx, column=15).fill = yellow_fill

            for col in range(1, 16):
                max_length = 0
                column_letter = get_column_letter(col)
                for row in range(1, len(self.all_parts) + 2):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 45)

            ws.freeze_panes = 'A2'

            default_name = os.path.splitext(os.path.basename(self.asm_path))[0] + "_детали.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel файлы", "*.xlsx")],
                initialfile=default_name
            )

            if file_path:
                wb.save(file_path)
                messagebox.showinfo("Готово", f"Экспортировано {len(self.all_parts)} деталей\n{file_path}")

        except ImportError:
            messagebox.showerror("Ошибка", "Не установлен openpyxl. Установите: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка экспорта: {str(e)}")


if __name__ == "__main__":
    app = KompasSyncApp()
    app.mainloop()